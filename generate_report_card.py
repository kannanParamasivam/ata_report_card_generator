import openpyxl
from openpyxl.drawing.image import Image
from copy import copy

workbook_name = '2023-24 Nilai-7 3rd Trimester Report Card'

# Iterate through all rows and columns in the sheet
student_name_row = 3
first_student_name_col = 6
total_row = 38
title_total_row = 2
comment_row = total_row + 3
question_col = 2
right_margin_col = 'P'

# Iterate through student names
def copy_image(source_sheet, new_sheet):
    for img in source_sheet._images:
        new_sheet.add_image(copy(img))

def delete_other_student_cols(cell, new_sheet):
    cols_to_delete_indexes = []

    for col_index in range(first_student_name_col, new_sheet.max_column + 1):
        
        if new_sheet.cell(student_name_row, col_index).value is None:
            new_sheet.column_dimensions[right_margin_col].width=3
            break;
        
        if col_index is not cell.col_idx:
            cols_to_delete_indexes.append(col_index)
                
    for col_index_del in reversed(cols_to_delete_indexes):
        new_sheet.delete_cols(col_index_del)

def main():

    # Open the Excel workbook
    workbook = openpyxl.load_workbook(f'{workbook_name}.xlsx')

    # Choose the sheet you want to iterate through
    source_sheet = workbook.worksheets[0]  # Replace 'Sheet1' with the name of your sheet

    for row in source_sheet.iter_rows(min_row=student_name_row, max_row=student_name_row, min_col=first_student_name_col):
        
        for cell in row:
        
            student_name = cell.value
            
            # terminate if student name is empty
            if student_name is None:
                break;
            
            elif student_name is not None:
                print(f'processing {student_name}')
                new_sheet = workbook.copy_worksheet(source_sheet)
                new_sheet.title = student_name

                new_sheet.cell(total_row, cell.col_idx).value = f"='{source_sheet.title}'!{source_sheet.cell(total_row, cell.col_idx).coordinate}"
                new_sheet.cell(title_total_row, cell.col_idx).value = f"='{source_sheet.title}'!{source_sheet.cell(title_total_row, cell.col_idx).coordinate}"
                command_ref_formula = f"'{source_sheet.title}'!{source_sheet.cell(comment_row, cell.col_idx).coordinate}"
                new_sheet.cell(comment_row+1, question_col).value = f"=IF({command_ref_formula}=\"\",\"\",{command_ref_formula})"
                new_sheet.cell(comment_row, cell.col_idx).value = None

                delete_other_student_cols(cell, new_sheet)

                copy_image(source_sheet, new_sheet)
                
                
            new_sheet.merge_cells(f'{new_sheet.cell(comment_row-1, question_col).coordinate}:F{comment_row-1}')
            new_sheet.merge_cells(f'{new_sheet.cell(comment_row-1, question_col).coordinate}:F{comment_row-1}') 
            new_sheet.merge_cells(f'{new_sheet.cell(comment_row-1, question_col).coordinate}:F{comment_row-1}') 
            new_sheet.merge_cells(new_sheet.cell(comment_row, question_col).coordinate + ':' + new_sheet.cell(comment_row, first_student_name_col).coordinate) 
            new_sheet.merge_cells(new_sheet.cell(comment_row, question_col).coordinate + ':' + new_sheet.cell(comment_row, first_student_name_col).coordinate) 
            new_sheet.merge_cells(new_sheet.cell(comment_row, question_col).coordinate + ':' + new_sheet.cell(comment_row, first_student_name_col).coordinate) 
            new_sheet.merge_cells(f'{new_sheet.cell(comment_row+1, question_col).coordinate}:F{comment_row+1}')
            new_sheet.merge_cells(f'{new_sheet.cell(comment_row+1, question_col).coordinate}:F{comment_row+1}')
            new_sheet.merge_cells(f'{new_sheet.cell(comment_row+1, question_col).coordinate}:F{comment_row+1}')
            
    
    workbook.save(f'{workbook_name}_splitted.xlsx')
    workbook.close()


if __name__ == "__main__":
    main()


